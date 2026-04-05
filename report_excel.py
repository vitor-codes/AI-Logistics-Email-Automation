"""Exporta o relatório em Excel com formatação profissional (apresentação / portfólio)."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
TITULO_PRINCIPAL = "Programações de entrega"
SUBTITULO_SISTEMA = "AI Logistics · Email Automation"


def exportar_relatorio_profissional(df: pd.DataFrame, path: Path) -> None:
    """
    Gera .xlsx com faixa de título, metadados de geração, tabela com cabeçalho
    estilizado, linhas zebradas e nota de rodapé.

    Espera colunas: material, volume, data_prevista (sem data_datetime).
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    ncols = 3

    wb = Workbook()
    ws = wb.active
    ws.title = "Programações"

    thin = Side(style="thin", color="D9D9D9")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_zebra = PatternFill(fill_type="solid", fgColor="FFF5F5F5")
    fill_titulo = PatternFill(fill_type="solid", fgColor="FF1F4E79")
    fill_cabecalho = PatternFill(fill_type="solid", fgColor="FF4472C4")

    font_titulo = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
    font_meta = Font(name="Calibri", size=10, color="FF5A5A5A")
    font_cab = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    font_corpo = Font(name="Calibri", size=11, color="FF333333")
    font_rodape = Font(name="Calibri", size=9, italic=True, color="FF888888")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=TITULO_PRINCIPAL)
    c.font = font_titulo
    c.fill = fill_titulo
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 38

    agora = datetime.now().strftime("%d/%m/%Y às %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(
        row=2,
        column=1,
        value=f"Relatório gerado em {agora} · {SUBTITULO_SISTEMA}",
    )
    c.font = font_meta
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24

    ws.row_dimensions[3].height = 8

    header_row = 4
    labels = ["Material", "Volume", "Data / hora prevista"]
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = font_cab
        cell.fill = fill_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all

    data_start = header_row + 1
    for i, row in enumerate(df.itertuples(index=False)):
        r = data_start + i
        valores = [row.material, row.volume, row.data_prevista]
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font_corpo
            cell.border = border_all
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
            elif col == 3:
                cell.number_format = "@"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        if i % 2 == 1:
            for col in range(1, ncols + 1):
                ws.cell(row=r, column=col).fill = fill_zebra

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26

    n_rows = len(df)
    foot_row = data_start + n_rows + 1
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=ncols)
    rodape = ws.cell(
        row=foot_row,
        column=1,
        value=(
            "Extração automatizada a partir de e-mails (Gmail API) com estruturação via IA "
            "(OpenAI GPT-4o-mini + LangChain). Relatório adequado para revisão operacional e "
            "demonstração de automação logística."
        ),
    )
    rodape.font = font_rodape
    rodape.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[foot_row].height = 48

    ws.freeze_panes = f"A{data_start}"

    wb.save(path)
